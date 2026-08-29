import io
from datetime import datetime
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def generate_excel_report(company, orders, product_sales, date_label, summary_metrics):
    """
    Generates an Excel workbook (.xlsx) containing Summary and Detailed Orders sheets.
    Returns BytesIO object.
    """
    wb = openpyxl.Workbook()
    
    # Styles
    navy_header_fill = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
    gold_fill = PatternFill(start_color="D38C44", end_color="D38C44", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="16213E")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    # --- Sheet 1: Sales Summary ---
    ws_summary = wb.active
    ws_summary.title = "Sales Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.append([company.name.upper()])
    ws_summary.append(["Sales & Performance Report"])
    ws_summary.append([f"Period: {date_label}  |  Generated: {timezone.localtime().strftime('%d-%b-%Y %I:%M %p')}"])
    ws_summary.append([])
    
    ws_summary["A1"].font = title_font
    ws_summary["A2"].font = Font(name="Calibri", size=13, bold=True, color="D38C44")
    ws_summary["A3"].font = subtitle_font
    
    # KPI Summary Cards Table
    ws_summary.append(["METRIC", "VALUE"])
    for cell in ws_summary[5]:
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    kpis = [
        ("Total Sales Revenue", summary_metrics.get('total_sales', 0.0), True),
        ("Total Orders Count", summary_metrics.get('total_orders', 0), False),
        ("Cash Collections", summary_metrics.get('total_cash', 0.0), True),
        ("UPI Collections", summary_metrics.get('total_upi', 0.0), True),
        ("Average Order Value", summary_metrics.get('avg_order_value', 0.0), True),
    ]
    
    for label, val, is_currency in kpis:
        row_idx = ws_summary.max_row + 1
        ws_summary.append([label, val])
        ws_summary.cell(row=row_idx, column=1).font = normal_font
        val_cell = ws_summary.cell(row=row_idx, column=2)
        val_cell.font = bold_font
        if is_currency:
            val_cell.number_format = '₹#,##0.00'
        ws_summary.cell(row=row_idx, column=1).border = thin_border
        val_cell.border = thin_border
    
    ws_summary.append([])
    ws_summary.append([])
    
    # Product Sales Table
    prod_header_row = ws_summary.max_row + 1
    ws_summary.append(["PRODUCT NAME", "QUANTITY SOLD", "TOTAL REVENUE (₹)"])
    for cell in ws_summary[prod_header_row]:
        cell.font = header_font
        cell.fill = gold_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    total_qty_sold = 0
    total_prod_rev = Decimal('0.00')
    for p in product_sales:
        row_idx = ws_summary.max_row + 1
        qty = p.get('total_qty', 0) or 0
        rev = float(p.get('total_revenue', 0) or 0)
        total_qty_sold += qty
        total_prod_rev += Decimal(str(rev))
        
        ws_summary.append([p.get('item_name', ''), qty, rev])
        ws_summary.cell(row=row_idx, column=1).font = normal_font
        ws_summary.cell(row=row_idx, column=2).font = normal_font
        rev_cell = ws_summary.cell(row=row_idx, column=3)
        rev_cell.font = normal_font
        rev_cell.number_format = '₹#,##0.00'
        
        for c in range(1, 4):
            ws_summary.cell(row=row_idx, column=c).border = thin_border
            if row_idx % 2 == 0:
                ws_summary.cell(row=row_idx, column=c).fill = alt_row_fill
    
    # Product Total Row
    tot_row_idx = ws_summary.max_row + 1
    ws_summary.append(["TOTAL", total_qty_sold, float(total_prod_rev)])
    for c in range(1, 4):
        cell = ws_summary.cell(row=tot_row_idx, column=c)
        cell.font = bold_font
        cell.fill = light_gray_fill
        cell.border = thin_border
    ws_summary.cell(row=tot_row_idx, column=3).number_format = '₹#,##0.00'

    # Auto-fit columns for Summary
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # --- Sheet 2: Detailed Orders Log ---
    ws_orders = wb.create_sheet(title="Detailed Orders")
    ws_orders.views.sheetView[0].showGridLines = True
    
    headers = [
        "Order #", "Date & Time", "Table", "Items Summary", "Status",
        "Payment Mode", "Cash (₹)", "UPI (₹)", "Total Amount (₹)", "Billed By"
    ]
    ws_orders.append(headers)
    for cell in ws_orders[1]:
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    for o in orders:
        row_idx = ws_orders.max_row + 1
        items_summary = ", ".join([f"{item.quantity}x {item.item_name}" for item in o.items.all()])
        table_label = f"Table {o.table.table_number}" if o.table else "Direct Sale"
        date_str = timezone.localtime(o.created_at).strftime('%d-%b-%Y %I:%M %p')
        billed_by_str = o.billed_by.username if o.billed_by else "--"
        
        ws_orders.append([
            o.order_number,
            date_str,
            table_label,
            items_summary,
            o.get_status_display(),
            o.get_payment_method_display(),
            float(o.cash_amount),
            float(o.upi_amount),
            float(o.total_amount),
            billed_by_str
        ])
        
        for c in range(1, 11):
            cell = ws_orders.cell(row=row_idx, column=c)
            cell.font = bold_font if c == 1 else normal_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill
        
        ws_orders.cell(row=row_idx, column=7).number_format = '₹#,##0.00'
        ws_orders.cell(row=row_idx, column=8).number_format = '₹#,##0.00'
        ws_orders.cell(row=row_idx, column=9).number_format = '₹#,##0.00'
    
    # Auto-fit columns for Orders sheet
    for col in ws_orders.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_orders.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_pdf_report(company, orders, product_sales, date_label, summary_metrics):
    """
    Generates a high quality PDF report using ReportLab.
    Returns BytesIO object.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#16213E')
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#555555')
    )
    
    section_heading = ParagraphStyle(
        'SectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=11,
        leading=15,
        textColor=colors.HexColor('#D38C44'),
        spaceBefore=12,
        spaceAfter=5
    )
    
    table_text = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#333333')
    )
    
    table_header = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white
    )
    
    elements = []
    
    # 1. Header & Title Block
    elements.append(Paragraph(f"<b>{company.name.upper()}</b>", title_style))
    elements.append(Spacer(1, 2))
    elements.append(Paragraph(f"<b>Sales & Financial Performance Report</b> &nbsp;|&nbsp; Period: <b>{date_label}</b>", ParagraphStyle('ReportSub', parent=subtitle_style, fontSize=10, textColor=colors.HexColor('#D38C44'))))
    elements.append(Paragraph(f"Generated on: {timezone.localtime().strftime('%d %B %Y, %I:%M %p')}", subtitle_style))
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#D38C44'), spaceAfter=12))
    
    # 2. KPI Summary Boxes
    total_sales = summary_metrics.get('total_sales', 0.0)
    total_orders = summary_metrics.get('total_orders', 0)
    total_cash = summary_metrics.get('total_cash', 0.0)
    total_upi = summary_metrics.get('total_upi', 0.0)
    avg_order = summary_metrics.get('avg_order_value', 0.0)
    
    kpi_data = [
        [
            Paragraph(f"<b>TOTAL SALES</b><br/><font size=12 color='#D38C44'><b>&#8377;{total_sales:,.2f}</b></font>", table_text),
            Paragraph(f"<b>TOTAL ORDERS</b><br/><font size=12 color='#16213E'><b>{total_orders}</b></font>", table_text),
            Paragraph(f"<b>CASH REVENUE</b><br/><font size=12 color='#4ECB71'><b>&#8377;{total_cash:,.2f}</b></font>", table_text),
            Paragraph(f"<b>UPI REVENUE</b><br/><font size=12 color='#6366F1'><b>&#8377;{total_upi:,.2f}</b></font>", table_text),
            Paragraph(f"<b>AVG ORDER VALUE</b><br/><font size=12 color='#16213E'><b>&#8377;{avg_order:,.2f}</b></font>", table_text),
        ]
    ]
    
    kpi_table = RLTable(kpi_data, colWidths=[104, 104, 104, 104, 104])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8F9FA')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E8DCCB')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E8DCCB')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 10))
    
    # 3. Product Sales Table
    elements.append(Paragraph("Product Sales Breakdown", section_heading))
    
    prod_data = [[
        Paragraph("Product Name", table_header),
        Paragraph("Quantity Sold", table_header),
        Paragraph("Revenue (&#8377;)", table_header)
    ]]
    
    tot_p_qty = 0
    tot_p_rev = Decimal('0.00')
    
    for p in product_sales:
        qty = p.get('total_qty', 0) or 0
        rev = float(p.get('total_revenue', 0) or 0)
        tot_p_qty += qty
        tot_p_rev += Decimal(str(rev))
        prod_data.append([
            Paragraph(p.get('item_name', ''), table_text),
            Paragraph(f"{qty} sold", table_text),
            Paragraph(f"&#8377;{rev:,.2f}", table_text)
        ])
    
    # Product Total Row
    prod_data.append([
        Paragraph("<b>TOTAL</b>", table_text),
        Paragraph(f"<b>{tot_p_qty} items</b>", table_text),
        Paragraph(f"<b>&#8377;{float(tot_p_rev):,.2f}</b>", table_text)
    ])
    
    prod_table = RLTable(prod_data, colWidths=[260, 130, 130])
    prod_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5EDE6')),
    ]
    
    for i in range(1, len(prod_data) - 1):
        if i % 2 == 0:
            prod_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FAFAFA')))
            
    prod_table.setStyle(TableStyle(prod_style))
    elements.append(prod_table)
    elements.append(Spacer(1, 10))
    
    # 4. Detailed Orders Table
    elements.append(Paragraph("Orders Log", section_heading))
    
    order_data = [[
        Paragraph("Order #", table_header),
        Paragraph("Date & Time", table_header),
        Paragraph("Table", table_header),
        Paragraph("Status", table_header),
        Paragraph("Payment", table_header),
        Paragraph("Amount (&#8377;)", table_header),
        Paragraph("Billed By", table_header)
    ]]
    
    for o in orders:
        table_label = f"T-{o.table.table_number}" if o.table else "Direct"
        date_str = timezone.localtime(o.created_at).strftime('%d-%b %I:%M%p')
        billed_str = o.billed_by.username if o.billed_by else "--"
        order_data.append([
            Paragraph(f"<b>{o.order_number}</b>", table_text),
            Paragraph(date_str, table_text),
            Paragraph(table_label, table_text),
            Paragraph(o.get_status_display(), table_text),
            Paragraph(o.get_payment_method_display(), table_text),
            Paragraph(f"&#8377;{float(o.total_amount):,.2f}", table_text),
            Paragraph(billed_str, table_text)
        ])
    
    if len(orders) == 0:
        order_data.append([Paragraph("No orders found for this period.", table_text), "", "", "", "", "", ""])
        
    orders_table = RLTable(order_data, colWidths=[75, 85, 55, 65, 75, 85, 80])
    orders_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F3460')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
    ]
    for i in range(1, len(order_data)):
        if i % 2 == 0:
            orders_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FAFAFA')))
            
    orders_table.setStyle(TableStyle(orders_style))
    elements.append(orders_table)
    
    def add_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#888888'))
        footer_text = f"BITPOS Report  |  {company.name}  |  Page {doc.page}"
        canvas.drawRightString(A4[0] - 36, 20, footer_text)
        canvas.restoreState()
    
    doc.build(elements, onFirstPage=add_footer, onLaterPages=add_footer)
    buffer.seek(0)
    return buffer
